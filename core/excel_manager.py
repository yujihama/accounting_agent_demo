import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io
from typing import Dict, List, Any, Optional, Tuple
from datetime import datetime
import json

class ExcelManager:
    """
    調書（Excel）の読み書き・操作機能
    """
    
    def __init__(self):
        self.workbook = None
        self.file_name = None
        self.original_data = {}
    
    def load_workbook(self, uploaded_file) -> Dict[str, Any]:
        """
        Excelファイル（調書）を読み込み
        
        Args:
            uploaded_file: StreamlitのExcelファイルアップロード
            
        Returns:
            読み込み結果とワークブック情報
        """
        try:
            # ファイル内容を読み込み
            file_content = uploaded_file.read()
            uploaded_file.seek(0)  # ポインタリセット
            
            # openpyxlでワークブックを開く
            self.workbook = openpyxl.load_workbook(io.BytesIO(file_content))
            self.file_name = uploaded_file.name
            
            # 全シートの情報を取得
            sheet_info = {}
            for sheet_name in self.workbook.sheetnames:
                sheet = self.workbook[sheet_name]
                sheet_info[sheet_name] = {
                    "max_row": sheet.max_row,
                    "max_column": sheet.max_column,
                    "has_data": sheet.max_row > 1 or sheet.max_column > 1
                }
                
                # シートの一部データをプレビュー用に保存
                if sheet_info[sheet_name]["has_data"]:
                    preview_data = []
                    for row in range(1, min(6, sheet.max_row + 1)):  # 最初の5行
                        row_data = []
                        for col in range(1, min(6, sheet.max_column + 1)):  # 最初の5列
                            cell_value = sheet.cell(row=row, column=col).value
                            row_data.append(str(cell_value) if cell_value is not None else "")
                        preview_data.append(row_data)
                    sheet_info[sheet_name]["preview"] = preview_data
            
            return {
                "success": True,
                "file_name": self.file_name,
                "sheet_names": self.workbook.sheetnames,
                "sheet_info": sheet_info,
                "loaded_at": datetime.now().isoformat()
            }
            
        except Exception as e:
            return {
                "success": False,
                "error": f"Excelファイル読み込みエラー: {str(e)}"
            }
    
    def write_results(self, target_sheet: str, start_row: int,
                     column_definitions: Dict[str, Dict[str, str]], 
                     data: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        結果を指定シートに書き込み
        
        Args:
            target_sheet: 対象シート名
            start_row: 開始行番号
            column_definitions: 列ごとの定義
                例: {"A": {"key": "row_number", "header": "No."}, 
                     "B": {"key": "data_id", "header": "データID"}}
            data: 書き込むデータのリスト
            
        Returns:
            書き込み結果
        """
        try:
            if not self.workbook:
                return {"success": False, "error": "ワークブックが読み込まれていません"}
            
            # シートが存在しない場合は作成
            if target_sheet not in self.workbook.sheetnames:
                self.workbook.create_sheet(target_sheet)
            
            sheet = self.workbook[target_sheet]
            
            # データを書き込み
            return self._write_results_to_sheet(sheet, start_row, column_definitions, data)
            
        except Exception as e:
            return {
                "success": False,
                "error": f"結果書き込みエラー: {str(e)}"
            }
    
    def _write_results_to_sheet(self, sheet, start_row: int, 
                               column_definitions: Dict[str, Dict[str, str]], 
                               data: List[Dict[str, Any]]) -> Dict[str, Any]:
        """データをシートに書き込み"""
        try:
            # ヘッダー行を書き込み
            for col_letter, col_def in sorted(column_definitions.items()):
                col_index = column_index_from_string(col_letter)
                cell = sheet.cell(row=start_row, column=col_index)
                cell.value = col_def["header"]
                # ヘッダーのスタイルを設定
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # データ行を書き込み
            for row_idx, row_data in enumerate(data):
                current_row = start_row + 1 + row_idx
                
                for col_letter, col_def in sorted(column_definitions.items()):
                    col_index = column_index_from_string(col_letter)
                    cell = sheet.cell(row=current_row, column=col_index)
                    
                    # 特別な処理：row_numberの場合は行番号を自動設定
                    if col_def["key"] == "row_number":
                        cell.value = row_idx + 1
                    else:
                        cell.value = row_data.get(col_def["key"], "")
                    
                    # データのスタイルを設定
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    
                    # 条件付きスタイル（例：エラーの場合は赤背景）
                    if col_def["key"] == "match_status" and "不一致" in str(cell.value):
                        cell.fill = PatternFill(start_color="FFB3B3", end_color="FFB3B3", fill_type="solid")
            
            # 書き込み情報
            written_rows = len(data)
            end_row = start_row + written_rows
            
            # 最終列を取得
            last_col_letter = max(column_definitions.keys())
            
            return {
                "success": True,
                "written_rows": written_rows,
                "range": f"A{start_row}:{last_col_letter}{end_row}",
                "target_sheet": sheet.title
            }
            
        except Exception as e:
            return {
                "success": False,
                "error": f"書き込みエラー: {str(e)}"
            }
    
    def save_workbook(self) -> bytes:
        """
        編集されたワークブックをバイトデータとして保存
        
        Returns:
            Excelファイルのバイトデータ
        """
        if not self.workbook:
            raise ValueError("ワークブックが読み込まれていません")
        
        output = io.BytesIO()
        self.workbook.save(output)
        output.seek(0)
        return output.getvalue()
    
    def read_sheet_range(self, sheet_name: str, start_cell: str, end_cell: str) -> List[List[str]]:
        """
        指定シートの指定範囲を読み取り
        
        Args:
            sheet_name: シート名
            start_cell: 開始セル
            end_cell: 終了セル
            
        Returns:
            セル値の2次元リスト
        """
        if not self.workbook or sheet_name not in self.workbook.sheetnames:
            return []
        
        sheet = self.workbook[sheet_name]
        
        try:
            start_row, start_col = self._parse_cell_reference(start_cell)
            end_row, end_col = self._parse_cell_reference(end_cell)
            
            data = []
            for row in range(start_row, end_row + 1):
                row_data = []
                for col in range(start_col, end_col + 1):
                    cell_value = sheet.cell(row=row, column=col).value
                    row_data.append(str(cell_value) if cell_value is not None else "")
                data.append(row_data)
            
            return data
            
        except Exception as e:
            print(f"範囲読み取りエラー: {str(e)}")
            return []
    
    def create_output_config_ui(self) -> Dict[str, Any]:
        """
        出力設定用のUI情報を生成
        
        Returns:
            UI生成用の情報
        """
        if not self.workbook:
            return {"error": "ワークブックが読み込まれていません"}
        
        return {
            "available_sheets": self.workbook.sheetnames,
            "current_file": self.file_name,
            "sample_cells": ["A1", "B1", "A3", "B3", "C3"],
            "column_samples": [
                {"key": "data_id", "header": "データID"},
                {"key": "invoice_amount", "header": "請求金額"},
                {"key": "payment_amount", "header": "入金金額"},
                {"key": "match_status", "header": "照合結果"},
                {"key": "variance", "header": "差額"},
                {"key": "remarks", "header": "備考"}
            ]
        }
    
    def _parse_cell_reference(self, cell_ref: str) -> Tuple[int, int]:
        """
        セル参照（例: "B3"）を行・列番号に変換
        
        Returns:
            (row, column) - 1ベースの番号
        """
        from openpyxl.utils import column_index_from_string
        
        # 文字と数字を分離
        col_str = ""
        row_str = ""
        
        for char in cell_ref.upper():
            if char.isalpha():
                col_str += char
            elif char.isdigit():
                row_str += char
        
        row = int(row_str) if row_str else 1
        col = column_index_from_string(col_str) if col_str else 1
        
        return row, col
    
    def get_sheet_preview(self, sheet_name: str, max_rows: int = 10, max_cols: int = 10) -> List[List[str]]:
        """
        指定シートのプレビューデータを取得
        
        Args:
            sheet_name: シート名
            max_rows: 最大行数
            max_cols: 最大列数
            
        Returns:
            プレビューデータ
        """
        if not self.workbook or sheet_name not in self.workbook.sheetnames:
            return []
        
        sheet = self.workbook[sheet_name]
        preview_data = []
        
        for row in range(1, min(max_rows + 1, sheet.max_row + 1)):
            row_data = []
            for col in range(1, min(max_cols + 1, sheet.max_column + 1)):
                cell_value = sheet.cell(row=row, column=col).value
                row_data.append(str(cell_value) if cell_value is not None else "")
            preview_data.append(row_data)
        
        return preview_data
    
    def validate_excel_file(self, uploaded_file) -> Tuple[bool, Optional[str]]:
        """
        Excelファイルの妥当性を検証
        
        Returns:
            (is_valid, error_message)
        """
        try:
            # ファイルサイズチェック
            max_size = 50 * 1024 * 1024  # 50MB
            if uploaded_file.size > max_size:
                return False, f"ファイルサイズが大きすぎます（最大50MB）: {uploaded_file.size / 1024 / 1024:.1f}MB"
            
            # Excelファイルかチェック
            if not uploaded_file.name.lower().endswith(('.xlsx', '.xls')):
                return False, "Excelファイル（.xlsx または .xls）をアップロードしてください"
            
            # ファイルの読み取り可能性をチェック
            file_content = uploaded_file.read()
            uploaded_file.seek(0)  # ポインタリセット
            
            test_workbook = openpyxl.load_workbook(io.BytesIO(file_content))
            
            # 最低1つのシートがあるかチェック
            if not test_workbook.sheetnames:
                return False, "有効なシートが見つかりません"
            
            return True, None
            
        except Exception as e:
            return False, f"Excelファイル検証エラー: {str(e)}"